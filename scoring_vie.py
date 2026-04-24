import pandas as pd
from openpyxl import load_workbook

FILE = "base_offres_vie.xlsx"


AFRICA_CODES = {
    "DZ", "AO", "BJ", "BW", "BF", "BI", "CV", "CM", "CF", "TD", "KM", "CG", "CD",
    "CI", "DJ", "EG", "GQ", "ER", "SZ", "ET", "GA", "GM", "GH", "GN", "GW", "KE",
    "LS", "LR", "LY", "MG", "MW", "ML", "MR", "MU", "YT", "MA", "MZ", "NA", "NE",
    "NG", "RE", "RW", "SH", "ST", "SN", "SC", "SL", "SO", "ZA", "SS", "SD", "TZ",
    "TG", "TN", "UG", "EH", "ZM", "ZW"
}

AFRICA_COUNTRY_KEYWORDS = [
    "algerie", "algérie", "algeria",
    "angola",
    "benin", "bénin",
    "botswana",
    "burkina", "burkina faso",
    "burundi",
    "cap-vert", "cape verde",
    "cameroun", "cameroon",
    "republique centrafricaine", "central african republic",
    "tchad", "chad",
    "comores", "comoros",
    "congo", "republique du congo", "république du congo",
    "rdc", "republique democratique du congo", "république démocratique du congo",
    "cote d'ivoire", "côte d'ivoire", "ivory coast",
    "djibouti",
    "egypte", "égypte", "egypt",
    "guinee equatoriale", "guinée équatoriale", "equatorial guinea",
    "erythree", "érythrée", "eritrea",
    "eswatini", "swaziland",
    "ethiopie", "éthiopie", "ethiopia",
    "gabon",
    "gambie", "gambia",
    "ghana",
    "guinee", "guinée", "guinea",
    "guinee-bissau", "guinée-bissau",
    "kenya",
    "lesotho",
    "liberia",
    "libye", "libya",
    "madagascar",
    "malawi",
    "mali",
    "mauritanie", "mauritania",
    "maurice", "mauritius",
    "maroc", "morocco",
    "mozambique",
    "namibie", "namibia",
    "niger",
    "nigeria", "nigéria",
    "rwanda",
    "sao tome", "são tomé", "sao tome-et-principe",
    "senegal", "sénégal",
    "seychelles",
    "sierra leone",
    "somalie", "somalia",
    "afrique du sud", "south africa",
    "soudan", "sudan",
    "soudan du sud", "south sudan",
    "tanzanie", "tanzania",
    "togo",
    "tunisie", "tunisia",
    "ouganda", "uganda",
    "zambie", "zambia",
    "zimbabwe",
    "afrique"
]

INDIA_KEYWORDS = ["inde", "india"]
INDIA_CODES = {"IN"}


def clean_text(text):
    if pd.isna(text):
        return ""
    return str(text).strip().lower()


def is_valid_date(date_str):
    if pd.isna(date_str):
        return False
    try:
        date = pd.to_datetime(date_str)
        return date >= pd.Timestamp("2026-09-01")
    except Exception:
        return False


def is_excluded_geo(country, country_code, city, description):
    country_text = clean_text(country)
    code_text = clean_text(country_code).upper()
    city_text = clean_text(city)
    desc_text = clean_text(description)

    combined = f"{country_text} {city_text} {desc_text}"

    if code_text in INDIA_CODES:
        return True, "Inde"
    if code_text in AFRICA_CODES:
        return True, "Afrique"

    if any(keyword in combined for keyword in INDIA_KEYWORDS):
        return True, "Inde"

    if any(keyword in combined for keyword in AFRICA_COUNTRY_KEYWORDS):
        return True, "Afrique"

    return False, ""


def score_destination(text):
    text = clean_text(text)

    priorite_1 = [
        "seoul", "corée du sud", "coree du sud", "south korea", "korea",
        "tokyo", "japan", "japon", "osaka",
        "china", "chine", "shanghai", "beijing", "pekin", "pékin", "hong kong",
        "singapore"
    ]

    priorite_2 = [
        "usa", "united states", "etats-unis", "états-unis",
        "new york", "san francisco", "boston", "chicago", "los angeles",
        "london", "londres",
        "canada", "toronto", "vancouver", "montreal", "montréal"
    ]

    grandes_villes = [
        "new york", "san francisco", "boston", "chicago", "los angeles",
        "london", "toronto", "vancouver", "montreal", "montréal",
        "seoul", "tokyo", "osaka", "shanghai", "beijing", "pékin", "singapore", "hong kong",
        "taipei", "sydney", "melbourne", "berlin", "madrid", "barcelona", "barcelone",
        "amsterdam", "dublin", "brussels", "bruxelles", "paris"
    ]

    if any(x in text for x in priorite_1):
        return 6, "Priorité 1"

    if any(x in text for x in priorite_2):
        return 4, "Priorité 2"

    if any(x in text for x in grandes_villes):
        return 2, "Grande ville dynamique"

    return 0, "Autre"


def score_poste(text):
    text = clean_text(text)
    score = 0

    mots_forts = [
        "analyst", "analysis",
        "business",
        "operations", "ops",
        "project",
        "strategy", "strategic",
        "finance", "financial",
        "reporting",
        "performance",
        "coordination",
        "transformation"
    ]

    mots_moyens = [
        "sales",
        "support",
        "planning",
        "crm",
        "assistant",
        "consulting",
        "program", "programme",
        "process",
        "marketing"
    ]

    # RH / HR explicitement pénalisé
    mots_negatifs = [
        "human resources", "hr", "rh", "talent acquisition", "recruitment", "recrutement",
        "people operations", "people partner",
        "chemist", "chemistry", "chimie",
        "biology", "biologie",
        "laboratory", "laboratoire", "lab",
        "pharmaceutical", "pharmaceutique",
        "petroleum engineering",
        "technician", "technicien",
        "maintenance",
        "mechanic", "mecanic", "mécanique",
        "nurse", "infirm"
        "marketing", "brand", "communication marketing"
    ]

    for mot in mots_forts:
        if mot in text:
            score += 2

    for mot in mots_moyens:
        if mot in text:
            score += 1

    for mot in mots_negatifs:
        if mot in text:
            score -= 4

    return max(0, min(score, 10))


def score_profil_recherche(text):
    text = clean_text(text)
    score = 0

    bonus = [
        "english", "anglais",
        "international",
        "analytical", "analytique",
        "communication",
        "problem solving", "resolution de problemes", "résolution de problèmes",
        "structured", "structure", "structuré", "structurée",
        "organized", "organise", "organisé", "organisée",
        "project management", "gestion de projet",
        "coordination",
        "business",
        "finance",
        "operations",
        "autonomy", "autonomie"
    ]

    malus = [
        "human resources", "hr", "rh", "talent acquisition", "recruitment", "recrutement",
        "people operations", "people partner",
        "marketing", "digital marketing", "brand", "communication marketing", "content marketing",
        "human resources", "hr", "rh", "talent acquisition", "recruitment", "recrutement",
        "chemistry", "chimie",
        "biology", "biologie",
        "biochemistry", "biochimie",
        "pharmaceutical", "pharmaceutique",
        "petroleum engineering",
        "laboratory", "laboratoire", "lab",
        "formulation",
        "technician", "technicien",
        "maintenance",
        "mechanic", "mecanic", "mécanique",
        "nurse", "infirm"
        
    ]

    for mot in bonus:
        if mot in text:
            score += 1

    for mot in malus:
        if mot in text:
            score -= 2

    return max(0, min(score, 4))


def classify(score):
    if score >= 16:
        return "Top priorité"
    if score >= 13:
        return "Très intéressant"
    if score >= 10:
        return "À regarder"
    if score >= 6:
        return "Secondaire"
    return "Faible intérêt"


def autofit_sheet_columns(file_path, sheet_name):
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.save(file_path)
        return

    ws = wb[sheet_name]

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except Exception:
                pass

        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)


def main():
    print("Chargement du fichier...")

    xls = pd.ExcelFile(FILE, engine="openpyxl")
    sheet_names = xls.sheet_names

    if not sheet_names:
        raise ValueError("Aucun onglet trouvé dans le fichier Excel.")

    source_sheet = "Base" if "Base" in sheet_names else sheet_names[0]
    print(f"Onglet source utilisé : {source_sheet}")

    df = pd.read_excel(FILE, sheet_name=source_sheet, engine="openpyxl")

    score_poste_list = []
    score_destination_list = []
    score_profil_list = []
    score_total_list = []
    priorite_list = []
    valide_date_list = []
    niveau_destination_list = []
    motif_exclusion_list = []

    for _, row in df.iterrows():
        poste = clean_text(row.get("poste"))
        ville = clean_text(row.get("ville"))
        pays = clean_text(row.get("pays"))
        pays_code = row.get("pays_code")
        description = clean_text(row.get("description_mission"))
        profil = clean_text(row.get("profil_recherche"))
        date_debut = row.get("date_debut")

        if not is_valid_date(date_debut):
            score_poste_list.append(0)
            score_destination_list.append(0)
            score_profil_list.append(0)
            score_total_list.append(-999)
            priorite_list.append("Hors cible (date)")
            valide_date_list.append(False)
            niveau_destination_list.append("Hors cible")
            motif_exclusion_list.append("Date de début avant 2026-09-01")
            continue

        excluded, motif = is_excluded_geo(
            country=pays,
            country_code=pays_code,
            city=ville,
            description=description
        )

        if excluded:
            score_poste_list.append(0)
            score_destination_list.append(0)
            score_profil_list.append(0)
            score_total_list.append(-999)
            priorite_list.append("Hors cible (geo)")
            valide_date_list.append(False)
            niveau_destination_list.append("Exclu")
            motif_exclusion_list.append(motif)
            continue

        texte_destination = f"{ville} {pays} {description}"
        texte_poste = f"{poste} {description}"
        texte_profil = profil

        s_dest, niveau_dest = score_destination(texte_destination)
        s_poste = score_poste(texte_poste)
        s_profil = score_profil_recherche(texte_profil)

        total = s_poste + s_dest + s_profil

        score_poste_list.append(s_poste)
        score_destination_list.append(s_dest)
        score_profil_list.append(s_profil)
        score_total_list.append(total)
        priorite_list.append(classify(total))
        valide_date_list.append(True)
        niveau_destination_list.append(niveau_dest)
        motif_exclusion_list.append("")

    df_scored = df.copy()
    df_scored["score_poste"] = score_poste_list
    df_scored["score_destination"] = score_destination_list
    df_scored["score_profil_recherche"] = score_profil_list
    df_scored["score_total"] = score_total_list
    df_scored["priorite"] = priorite_list
    df_scored["valide_date"] = valide_date_list
    df_scored["niveau_destination"] = niveau_destination_list
    df_scored["motif_exclusion"] = motif_exclusion_list

    top_priorite = df_scored[
        (df_scored["valide_date"] == True) &
        (df_scored["score_total"] >= 10)
    ].copy()

    top_priorite = top_priorite.sort_values(
        by=["score_total", "score_poste", "score_destination", "score_profil_recherche"],
        ascending=False
    )

    print("Écriture dans le fichier...")

    with pd.ExcelWriter(FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        top_priorite.to_excel(writer, sheet_name="Top priorite", index=False)

    autofit_sheet_columns(FILE, "Top priorite")

    print("Terminé. Onglet 'Top priorite' créé / mis à jour.")
    print("Barème : poste /10, destination /6, profil recherché /4, total /20.")


if __name__ == "__main__":
    main()
