import pandas as pd
from openpyxl import load_workbook
import os

FILE = "base_offres_vie.xlsx"
TOP_SHEET = "Top priorite"
WORK_SHEET = "Travail"

WORK_COLUMNS = [
    "id",
    "lien_offre",
    "Poste",
    "Date_debut",
    "Pays",
    "Ville",
    "Entreprise",
    "Statut",
    "Priorité",
    "Date candidature",
    "Score_total",
    "Niveau_destination",
    "nouvelle_offre",
    "active_ce_jour"
]


def load_sheet_if_exists(file_path, sheet_name):
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    if sheet_name in xls.sheet_names:
        return pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
    return pd.DataFrame()


def build_new_work_rows(top_df):
    rows = pd.DataFrame()

    rows["id"] = top_df["id"]
    rows["lien_offre"] = top_df["lien_offre"]
    rows["Poste"] = top_df["poste"]
    rows["Date_debut"] = top_df["date_debut"]
    rows["Pays"] = top_df["pays"]
    rows["Ville"] = top_df["ville"]
    rows["Entreprise"] = top_df["entreprise"]
    rows["Statut"] = ""
    rows["Priorité"] = top_df["priorite"]
    rows["Date candidature"] = ""
    rows["Score_total"] = top_df["score_total"]
    rows["Niveau_destination"] = top_df["niveau_destination"]
    rows["nouvelle_offre"] = top_df["nouvelle_offre"]
    rows["active_ce_jour"] = top_df["active_ce_jour"]

    return rows


def ensure_work_columns(df):
    for col in WORK_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    return df[WORK_COLUMNS].copy()


def autofit_sheet_columns(file_path, sheet_name):
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.save(file_path)
        return

    ws = wb[sheet_name]

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except Exception:
                pass

        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)


def main():
    if not os.path.exists(FILE):
        raise FileNotFoundError(f"Fichier introuvable : {FILE}")

    print("Lecture de l'onglet Top priorite...")
    top_df = load_sheet_if_exists(FILE, TOP_SHEET)

    if top_df.empty:
        raise ValueError("L'onglet 'Top priorite' est vide ou introuvable.")

    needed_in_top = [
        "id", "lien_offre", "poste", "date_debut", "pays", "ville",
        "entreprise", "priorite", "score_total",
        "niveau_destination", "nouvelle_offre", "active_ce_jour"
    ]
    missing = [c for c in needed_in_top if c not in top_df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes dans 'Top priorite' : {missing}")

    new_rows = build_new_work_rows(top_df)
    new_rows = ensure_work_columns(new_rows)

    print("Lecture de l'onglet Travail...")
    work_df = load_sheet_if_exists(FILE, WORK_SHEET)

    if work_df.empty:
        print("Aucun onglet Travail existant, création d'un nouveau.")
        final_work_df = new_rows.copy()
        added_count = len(final_work_df)
    else:
        work_df = ensure_work_columns(work_df)

        work_df["id"] = pd.to_numeric(work_df["id"], errors="coerce")
        new_rows["id"] = pd.to_numeric(new_rows["id"], errors="coerce")

        existing_ids = set(work_df["id"].dropna().astype(int).tolist())

        rows_to_add = new_rows[~new_rows["id"].isin(existing_ids)].copy()
        added_count = len(rows_to_add)

        final_work_df = pd.concat([work_df, rows_to_add], ignore_index=True)

    final_work_df["nouvelle_offre"] = final_work_df["nouvelle_offre"].fillna(False)
    final_work_df["Score_total"] = pd.to_numeric(final_work_df["Score_total"], errors="coerce").fillna(-999)

    final_work_df = final_work_df.sort_values(
        by=["nouvelle_offre", "Score_total"],
        ascending=[False, False]
    ).reset_index(drop=True)

    print("Écriture de l'onglet Travail...")

    with pd.ExcelWriter(FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        final_work_df.to_excel(writer, sheet_name=WORK_SHEET, index=False)

    autofit_sheet_columns(FILE, WORK_SHEET)

    print(f"Terminé. {added_count} nouvelle(s) ligne(s) ajoutée(s) dans l'onglet Travail.")
    print("Aucune ligne existante n'a été supprimée.")


if __name__ == "__main__":
    main()
